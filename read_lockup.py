import openpyxl

wb = openpyxl.load_workbook(r'c:\Users\mark\Desktop\raq.store\lockup-html.xlsx')
ws = wb.active

print(f"工作表: {ws.title}")
print(f"总行数: {ws.max_row}")
print(f"\n{'序号':<6} {'地址':<44} {'锁仓数量':>12}")
print("-" * 65)

for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
    address = row[0].value
    amount = row[1].value
    print(f"{row_num:<6} {address:<44} {amount:>12,}")

print("-" * 65)

total = sum((row[0].value if isinstance(row, tuple) else row.value) or 0 for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2, values_only=False))
total = sum(cell.value for cell in ws['B'])
print(f"\n总计锁仓数量: {total:,}")
